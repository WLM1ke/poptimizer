"""Загрузка данных о потребительской инфляции."""
import io
import logging
import types
from datetime import datetime, timedelta
from typing import ClassVar, Final

import aiohttp
from openpyxl.reader import excel
from openpyxl.worksheet import worksheet
from pydantic import Field, ValidationError, validator

from poptimizer.data import domain, exceptions
from poptimizer.data.repo import Repo

_URL: Final = "https://rosstat.gov.ru/storage/mediabank/ipc_4(2).xlsx"
_SHEET_NAME: Final = "01"

_FIRST_MONTH_CELL: Final = "A6"
_FIRST_MONTH_VALUE: Final = "январь"
_FIRST_YEAR_CELL: Final = "B4"
_FIRST_YEAR_VALUE: Final = 1991
_DATA_RANGE: Final = types.MappingProxyType(
    {
        "min_row": 6,
        "max_row": 17,
        "min_col": 2,
        "values_only": True,
    },
)

_JANUARY_LAST_DAY: Final = 31

_MINIMUM_MONTHLY_CPI: Final = 0.99


class CPI(domain.Row):
    """Значение мультипликативной инфляции для заданного месяца (его последней даты)."""

    date: datetime
    cpi: float = Field(gt=_MINIMUM_MONTHLY_CPI)

    @validator("date")
    def _must_be_last_day_of_month(cls, date: datetime) -> datetime:
        if (date + timedelta(days=1)).month == date.month:
            raise ValueError("not last day of month")

        return date


class Table(domain.Table):
    """Таблица с инфляцией."""

    group: ClassVar[domain.Group] = domain.Group.CPI
    df: list[CPI] = Field(default_factory=list[CPI])

    _must_be_sorted_by_date = validator("df", allow_reuse=True)(domain.validate_sorted_by_date)


class Service:
    """Сервис загрузки потребительской инфляции."""

    def __init__(self, repo: Repo, session: aiohttp.ClientSession) -> None:
        self._logger = logging.getLogger("CPISrv")
        self._repo = repo
        self._session = session

    async def update(self, update_day: datetime) -> None:
        """Обновляет потребительскую инфляцию и логирует неудачную попытку."""
        try:
            await self._update(update_day)
        except (aiohttp.ClientError, ValidationError, exceptions.DataError) as err:
            self._logger.warning(f"can't complete CPI update {err}")

            return

        self._logger.info("update is completed")

    async def _update(self, update_day: datetime) -> None:
        table = await self._repo.get(Table)
        table.timestamp = update_day

        xlsx_file = await self._download()
        df = _prepare_df(xlsx_file)

        if table.df != df[: len(table.df)]:
            raise exceptions.UpdateError("new cpi mismatch old")

        table.df = df

        await self._repo.save(table)

    async def _download(self) -> io.BytesIO:
        async with self._session.get(_URL) as resp:
            if not resp.ok:
                raise exceptions.UpdateError(f"bad CPI respond status {resp.reason}")

            return io.BytesIO(await resp.read())


def _prepare_df(xlsx: io.BytesIO) -> list[CPI]:
    ws = excel.load_workbook(xlsx)[_SHEET_NAME]

    _validate_data_position(ws)

    date = datetime(year=_FIRST_YEAR_VALUE, month=1, day=_JANUARY_LAST_DAY)
    rows: list[CPI] = []

    for row in ws.iter_cols(**_DATA_RANGE):
        for cell in row:
            if cell is None:
                return rows

            rows.append(CPI(date=date, cpi=cell / 100))

            date = _get_next_month_end(date)

    return rows


def _validate_data_position(ws: worksheet.Worksheet) -> None:
    if (first_month := ws[_FIRST_MONTH_CELL].value) != _FIRST_MONTH_VALUE:
        raise exceptions.UpdateError(f"wrong first month {first_month}")
    if (first_year := ws[_FIRST_YEAR_CELL].value) != _FIRST_YEAR_VALUE:
        raise exceptions.UpdateError(f"first year {first_year}")


def _get_next_month_end(date: datetime) -> datetime:
    skip_month = date + timedelta(days=_JANUARY_LAST_DAY + 1)

    return skip_month - timedelta(days=skip_month.day)
