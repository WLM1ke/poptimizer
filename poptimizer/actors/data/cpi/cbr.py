import io
from datetime import date, datetime, timedelta
from typing import TYPE_CHECKING, Final, cast

import aiohttp
from openpyxl.reader import excel

from poptimizer.actors.data.cpi import cpi
from poptimizer.adapters import http
from poptimizer.core import errors

if TYPE_CHECKING:
    import aiohttp
    from openpyxl.worksheet import worksheet


_URL: Final = "https://www.cbr.ru/Content/Document/File/108632/indicators_cpd.xlsx"

_SHEET_NAME: Final = "Лист1"
_FIRST_DATE_CELL: Final = "B1"
_FIRST_DATE_VALUE: Final = date(year=2002, month=1, day=1)
_VALUE_HEADER_CELL: Final = "A2"
_VALUE_HEADER_VALUE: Final = "Все товары и услуги"

_MIN_ROW: Final = 1
_MAX_ROW: Final = 2
_MIN_COL: Final = 2

_MAX_MONTH_DAYS: Final = 31


class Client:
    def __init__(self, http_client: aiohttp.ClientSession) -> None:
        self._http_client = http_client

    async def download_cpi(self) -> list[cpi.CPIRow]:
        async with (
            http.wrap_err("can't download CPI data"),
            self._http_client.get(_URL) as resp,
        ):
            if not resp.ok:
                raise errors.AdapterError(f"bad CPI respond status {resp.reason}")

            return _parse_rows(io.BytesIO(await resp.read()))


def _parse_rows(xlsx: io.BytesIO) -> list[cpi.CPIRow]:
    wb = excel.load_workbook(xlsx)
    ws = cast("worksheet.Worksheet", wb[_SHEET_NAME])

    _validate_data_position(ws)

    rows: list[cpi.CPIRow] = []

    for row in ws.iter_cols(min_row=_MIN_ROW, max_row=_MAX_ROW, min_col=_MIN_COL, values_only=True):
        day, value = row

        rows.append(cpi.CPIRow(day=_month_end(cast("datetime", day).date()), cpi=1 + cast("float", value) / 100))

    return rows


def _validate_data_position(ws: worksheet.Worksheet) -> None:
    if (first_date := ws[_FIRST_DATE_CELL].value.date()) != _FIRST_DATE_VALUE:
        raise errors.AdapterError(f"first date {first_date}")
    if (header := ws[_VALUE_HEADER_CELL].value) != _VALUE_HEADER_VALUE:
        raise errors.AdapterError(f"wrong header {header}")


def _month_end(day: date) -> date:
    skip_month = day.replace(day=1) + timedelta(days=_MAX_MONTH_DAYS)

    return skip_month - timedelta(days=skip_month.day)
