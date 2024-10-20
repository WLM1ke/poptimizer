import io
import re
from datetime import date, timedelta
from typing import Final, cast

import aiohttp
from openpyxl.reader import excel
from openpyxl.worksheet import worksheet

from poptimizer.domain import consts
from poptimizer.domain.entity_ import entity
from poptimizer.domain.entity_.data import cpi
from poptimizer.domain.service import domain_service

_PRICES_PAGE: Final = "https://rosstat.gov.ru/statistics/price"
_RE_FILE: Final = re.compile(r"/[iI]pc[\-_]mes[\-_][0-9]{1,2}-[0-9]{4}.xlsx")
_URL_TMPL: Final = "https://rosstat.gov.ru/storage/mediabank/{}"
_SHEET_NAME: Final = "01"

_FIRST_MONTH_CELL: Final = "A6"
_FIRST_MONTH_VALUE: Final = "январь"
_FIRST_YEAR_CELL: Final = "B4"
_FIRST_YEAR_VALUE: Final = 1991

_MIN_ROW: Final = 6
_MAX_ROW: Final = 17
_MIN_COL: Final = 2

_JANUARY_LAST_DAY: Final = 31


class CPIUpdateService:
    def __init__(self, http_session: aiohttp.ClientSession) -> None:
        self._http_session = http_session

    async def __call__(self, ctx: domain_service.Ctx, update_day: entity.Day) -> None:
        table = await ctx.get_for_update(cpi.CPI)

        xlsx_file = await self._download()
        row = _parse_rows(xlsx_file)

        table.update(update_day, row)

    async def _download(self) -> io.BytesIO:
        async with self._http_session.get(_PRICES_PAGE) as resp:
            if not resp.ok:
                raise consts.DomainError(f"bad CPI respond status {resp.reason}")

            html = await resp.text()

        if (file_name := _RE_FILE.search(html)) is None:
            raise consts.DomainError("can't find file with CPI")

        cpi_url = _URL_TMPL.format(file_name.group(0))

        async with self._http_session.get(cpi_url) as resp:
            if not resp.ok:
                raise consts.DomainError(f"bad CPI respond status {resp.reason}")

            return io.BytesIO(await resp.read())


def _parse_rows(xlsx: io.BytesIO) -> list[cpi.Row]:
    wb = excel.load_workbook(xlsx)
    ws = cast(worksheet.Worksheet, wb[_SHEET_NAME])

    _validate_data_position(ws)

    day = date(year=_FIRST_YEAR_VALUE, month=1, day=_JANUARY_LAST_DAY)
    rows: list[cpi.Row] = []

    for row in ws.iter_cols(min_row=_MIN_ROW, max_row=_MAX_ROW, min_col=_MIN_COL, values_only=True):
        for cell in row:
            match cell:
                case float() | int():
                    rows.append(cpi.Row(day=day, cpi=cell / 100))
                case None:
                    return rows
                case _:
                    raise consts.DomainError(f"strange CPI value {cell}")

            day = _get_next_month_end(day)

    return rows


def _validate_data_position(ws: worksheet.Worksheet) -> None:
    if (first_year := ws[_FIRST_YEAR_CELL].value) != _FIRST_YEAR_VALUE:
        raise consts.DomainError(f"first year {first_year}")
    if (first_month := ws[_FIRST_MONTH_CELL].value) != _FIRST_MONTH_VALUE:
        raise consts.DomainError(f"wrong first month {first_month}")


def _get_next_month_end(day: date) -> date:
    skip_month = day + timedelta(days=_JANUARY_LAST_DAY + 1)

    return skip_month - timedelta(days=skip_month.day)
