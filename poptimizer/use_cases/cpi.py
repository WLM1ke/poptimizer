import io
from datetime import date, datetime, timedelta
from typing import Final, cast

import aiohttp
from openpyxl.reader import excel
from openpyxl.worksheet import worksheet

from poptimizer import errors
from poptimizer.domain import cpi
from poptimizer.use_cases import handler

_URL: Final = "https://www.cbr.ru/Content/Document/File/108632/indicators_cpd.xlsx"
_SHEET_NAME: Final = "Лист1"
_FIRST_DATE_CELL: Final = "B1"
_FIRST_DATE_VALUE: Final = date(year=2002, month=1, day=1)
_VALUE_HEADER_CELL: Final = "A2"
_VALUE_HEADER_VALUE: Final = "Все товары и услуги"

_MIN_ROW: Final = 1
_MAX_ROW: Final = 2
_MIN_COL: Final = 2

_MAX_MONTH_DAYS: Final = 31


class CPIHandler:
    def __init__(self, http_session: aiohttp.ClientSession) -> None:
        self._http_session = http_session

    async def __call__(self, ctx: handler.Ctx, msg: handler.NewDataPublished) -> None:
        table = await ctx.get_for_update(cpi.CPI)

        try:
            xlsx_file = await self._download()
        except (TimeoutError, aiohttp.ClientError) as err:
            raise errors.UseCasesError("CPI download") from err

        row = _parse_rows(xlsx_file)

        table.update(msg.day, row)

    async def _download(self) -> io.BytesIO:
        async with self._http_session.get(_URL) as resp:
            if not resp.ok:
                raise errors.UseCasesError(f"bad CPI respond status {resp.reason}")

            return io.BytesIO(await resp.read())


def _parse_rows(xlsx: io.BytesIO) -> list[cpi.Row]:
    wb = excel.load_workbook(xlsx)
    ws = cast("worksheet.Worksheet", wb[_SHEET_NAME])

    _validate_data_position(ws)

    rows: list[cpi.Row] = []

    for row in ws.iter_cols(min_row=_MIN_ROW, max_row=_MAX_ROW, min_col=_MIN_COL, values_only=True):
        day, value = row
        try:
            rows.append(cpi.Row(day=_month_end(cast("datetime", day).date()), cpi=1 + cast("float", value) / 100))
        except ValueError as err:
            raise errors.UseCasesError("bad CPI data") from err

    return rows


def _validate_data_position(ws: worksheet.Worksheet) -> None:
    if (first_date := ws[_FIRST_DATE_CELL].value.date()) != _FIRST_DATE_VALUE:
        raise errors.UseCasesError(f"first date {first_date}")
    if (header := ws[_VALUE_HEADER_CELL].value) != _VALUE_HEADER_VALUE:
        raise errors.UseCasesError(f"wrong header {header}")


def _month_end(day: date) -> date:
    skip_month = day.replace(day=1) + timedelta(days=_MAX_MONTH_DAYS)

    return skip_month - timedelta(days=skip_month.day)
