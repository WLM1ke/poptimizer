import io
from datetime import date, datetime, timedelta
from typing import TYPE_CHECKING, Final, cast

from openpyxl.reader import excel
from openpyxl.utils.cell import get_column_letter

from poptimizer.core import errors
from poptimizer.data.cpi import cpi

if TYPE_CHECKING:
    from openpyxl.worksheet import worksheet


_SHEET_NAME: Final = "Лист1"

_DATES_ROW: Final = 1
_FIRST_DATE_COL: Final = 3
_FIRST_DATE_VALUE: Final = date(year=2002, month=1, day=1)
_CPI_HEADER_VALUE: Final = "Все товары и услуги"

_MAX_MONTH_DAYS: Final = 31


def cpi_parser(xlsx: io.BytesIO) -> list[cpi.Row]:
    wb = excel.load_workbook(xlsx)
    ws = cast("worksheet.Worksheet", wb[_SHEET_NAME])

    _validate_data_position(ws)

    rows: list[cpi.Row] = []

    for row in ws.iter_cols(
        min_row=_DATES_ROW,
        max_row=_DATES_ROW + 1,
        min_col=_FIRST_DATE_COL,
        values_only=True,
    ):
        day, value = row

        rows.append(cpi.Row(day=_month_end(_as_date(day)), cpi=1 + cast("float", value) / 100))

    return rows


def _validate_data_position(ws: worksheet.Worksheet) -> None:
    first_date_cell = f"{get_column_letter(_FIRST_DATE_COL)}{_DATES_ROW}"
    if (first_date := _as_date(ws[first_date_cell].value)) != _FIRST_DATE_VALUE:
        raise errors.AdapterError(f"first date {first_date}")

    value_header_cell = f"{get_column_letter(_FIRST_DATE_COL - 1)}{_DATES_ROW + 1}"
    if (header := ws[value_header_cell].value) != _CPI_HEADER_VALUE:
        raise errors.AdapterError(f"wrong header {header}")


def _as_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()

    raise errors.AdapterError(f"unexpected date cell value {value!r}")


def _month_end(day: date) -> date:
    skip_month = day.replace(day=1) + timedelta(days=_MAX_MONTH_DAYS)

    return skip_month - timedelta(days=skip_month.day)
